# /// script
# requires-python = ">=3.9"
# dependencies = ["openpyxl", "pyyaml"]
# ///
"""Institution salary-scale importer — xlsx → rate-table.yaml (Stage 5 costing input).

WHY this exists (year-variance): an institution's salary scale (base rates, on-cost %) RISES
EVERY YEAR, so the numbers are INSTANCE data, not skill data. This skill ships the METHOD + this
PARSER; the numbers live in a rate-table.yaml the applicant REGENERATES each year by re-running this
on the current calculator. build_budget.py then looks rates up (see its `rate_ref`), so no rate is
ever hand-typed or hardcoded in the skill.

Parses the standard UTS "Salary Scales" calculator layout (also works for similar HEW/academic
scale sheets): section headers name a classification LEVEL, following rows are STEPS carrying a base
salary (col B), a with-on-cost figure (col C = base × on-cost), and optional casual hourly (col F,
incl. loading). It DERIVES on_cost_pct from C/B and casual_loading is read from the header text.

Emitted rate-table.yaml (consumed by build_budget.py --rates):
    institution: UTS
    year: 2027
    source: "UTS_Salary_Calculator_ARC_LP26.xlsx"
    on_cost_pct: 0.30            # DERIVED from col C / col B
    casual_loading: 0.25         # read from the 'including 25% loading' header
    hdr_stipend: 39000           # fixed full-time HDR/PhD stipend (0 on-cost)
    scales:
      "Level A - Associate Lecturer, Research Associate (Postdoc)":
        "Step 1": {base: 98452.89, casual_hourly: null}
        "Step 3": {base: 111708.37, min_for_phd: true}
        ...
      "Research Assistant - HEW Level 5":
        "s1051": {base: 87738.56, casual_hourly: 60.05}

用法:
    uv run import_uts_rates.py UTS_Salary_Calculator_ARC_LP26.xlsx -o rate-table-uts-2027.yaml
    uv run import_uts_rates.py --self-test

退出码: 0 = 解析出 ≥1 scale; 2 = 结构错误（打不开 / 无可识别 scale）。
"""
import argparse
import re
import sys

STEP_RE = re.compile(r"^(step\s*\d+|s\d{3,4})$", re.I)
# an annual academic/professional base salary floor — guards against a silent mis-parse if a future
# year's calculator shifts columns (e.g. col B holds a step index or an hourly rate, not the base).
# A "base" below this is NOT treated as a salary → the scale drops → build_table fails closed.
MIN_PLAUSIBLE_BASE = 20000
# a level header: has text in col A, no base in col B, and looks like a classification line
LEVEL_HINT = re.compile(r"(level\s+[a-e]\b|hew\s*level|research assistant|lecturer|professor|associate)", re.I)


class ImportError_(ValueError):
    pass


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_workbook(ws):
    scales, cur = {}, None
    year, on_cost_samples, casual_loading, hdr = None, [], None, None
    for row in ws.iter_rows():
        cells = {c.column_letter: c.value for c in row if c.value is not None}
        a = str(cells.get("A", "")).strip()
        # year from a "<YYYY> salary" header
        if year is None:
            for v in cells.values():
                m = re.search(r"\b(20\d{2})\s+salary", str(v), re.I)
                if m:
                    year = int(m.group(1))
        # casual loading from header text
        if casual_loading is None:
            for v in cells.values():
                m = re.search(r"including\s+(\d+)%\s+loading", str(v), re.I)
                if m:
                    casual_loading = int(m.group(1)) / 100.0
        # HDR stipend
        if hdr is None:
            for v in cells.values():
                m = re.search(r"\$([\d,]+)\s*UTS rate", str(v))
                if m:
                    hdr = float(m.group(1).replace(",", ""))
        base = _f(cells.get("B"))
        # STEP row: A matches a step token AND B is a PLAUSIBLE annual base (not a step index / hourly
        # rate — fail-closed against a column shift in a future year's calculator)
        if STEP_RE.match(a) and base is not None and base >= MIN_PLAUSIBLE_BASE and cur is not None:
            c_val = _f(cells.get("C"))
            casual = _f(cells.get("F"))
            if c_val and base:
                on_cost_samples.append(c_val / base - 1.0)
            entry = {"base": round(base, 2)}
            if casual is not None:
                entry["casual_hourly"] = round(casual, 2)
            # carry a min-for-PhD flag if a neighbouring note says so
            note = " ".join(str(v) for v in cells.values())
            if re.search(r"minimum level.*phd", note, re.I):
                entry["min_for_phd"] = True
            scales[cur][a.title() if a.lower().startswith("step") else a] = entry
        # LEVEL header: text in A, no numeric base, matches a classification hint
        elif a and base is None and LEVEL_HINT.search(a) and not STEP_RE.match(a):
            cur = re.sub(r"\s+", " ", a.replace("\n", " ")).strip()
            scales.setdefault(cur, {})
    scales = {k: v for k, v in scales.items() if v}  # drop empty
    on_cost = round(sum(on_cost_samples) / len(on_cost_samples), 4) if on_cost_samples else None
    return year, on_cost, casual_loading, hdr, scales


def build_table(path):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise ImportError_(f"无法打开 xlsx: {exc}")
    ws = wb.worksheets[0]
    year, on_cost, casual_loading, hdr, scales = parse_workbook(ws)
    if not scales:
        raise ImportError_("未识别到任何 salary scale（layout 可能变了 — 检查 level 头 + step 行）")
    inst = "UTS" if "uts" in path.lower() else "[TO SET]"
    return {"institution": inst, "year": year, "source": path.split("/")[-1],
            "on_cost_pct": on_cost, "casual_loading": casual_loading,
            "hdr_stipend": hdr, "scales": scales}


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", nargs="?")
    ap.add_argument("-o", "--out")
    ap.add_argument("--self-test", action="store_true")
    args = ap.parse_args()
    if args.self_test:
        return self_test()
    if not args.xlsx:
        ap.error("需要 xlsx 路径 (或 --self-test)")
    try:
        table = build_table(args.xlsx)
    except ImportError_ as exc:
        print(f"IMPORT ERROR (fail-closed): {exc}", file=sys.stderr)
        return 2
    n = sum(len(v) for v in table["scales"].values())
    print(f"parsed {len(table['scales'])} scale(s), {n} step(s); year={table['year']} "
          f"on_cost={table['on_cost_pct']} casual_loading={table['casual_loading']} hdr={table['hdr_stipend']}")
    for lvl, steps in table["scales"].items():
        print(f"  {lvl}: {', '.join(steps)}")
    if args.out:
        import yaml
        yaml.safe_dump(table, open(args.out, "w", encoding="utf-8"), sort_keys=False, allow_unicode=True)
        print(f"\nwrote {args.out}")
    return 0


def self_test():
    # synthesise a tiny worksheet mimicking the UTS layout
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "UTS SALARY SCALES commencing 2027"
    ws["E6"] = "$39,000 UTS rate*"
    ws["A10"] = "Level A - Research Associate (Postdoc)"
    ws["B11"] = "2027 salary"
    ws["F11"] = "Casual Hourly Rate including 25% loading"
    ws["A12"] = "Step 1"; ws["B12"] = 100000.0; ws["C12"] = 130000.0; ws["F12"] = 60.0
    ws["A14"] = "Step 3"; ws["B14"] = 111000.0; ws["C14"] = 144300.0; ws["F14"] = 66.0
    year, on_cost, loading, hdr, scales = parse_workbook(ws)
    assert year == 2027, year
    assert abs(on_cost - 0.30) < 1e-6, on_cost              # DERIVED from 130000/100000
    assert loading == 0.25, loading
    assert hdr == 39000.0, hdr
    lvl = "Level A - Research Associate (Postdoc)"
    assert lvl in scales, scales.keys()
    assert scales[lvl]["Step 1"]["base"] == 100000.0, scales[lvl]
    assert scales[lvl]["Step 1"]["casual_hourly"] == 60.0, scales[lvl]
    assert "Step 3" in scales[lvl], scales[lvl]

    # plausibility floor: a column shift where col B holds a step index (1,2) → NOT parsed as base
    ws2 = openpyxl.Workbook().active
    ws2["A10"] = "Level A - Research Fellow"; ws2["B11"] = "Step no."
    ws2["A12"] = "Step 1"; ws2["B12"] = 1; ws2["C12"] = 100000.0
    ws2["A13"] = "Step 2"; ws2["B13"] = 2; ws2["C13"] = 105000.0
    _, _, _, _, sc2 = parse_workbook(ws2)
    assert not sc2, f"implausible base (step index) must NOT parse — got {sc2}"

    print("self-test OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
